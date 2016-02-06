import os
import json
import openpyxl

os.chdir("rep_name")
wb = openpyxl.load_workbook("excel.xlsx")
sheet = wb.active

excel_dict = {}
excel_list = []


def checkbool(check_string):
	string_lower = check_string.lower()
	if string_lower == 'true':
		check_int = 1
	elif string_lower == 'false':
		check_int = 0
	else:
		check_int = 2
	return check_int
	

for i in range(3, sheet.max_row + 1):
	for j in range(1, sheet.max_column + 1):
		tempValue = sheet.cell(row = i, column = j).value
		if tempValue == None:
				tempValue = sheet.cell(row = 2, column = j).value
		try:
			tempValue = json.loads(tempValue)
		except:
			try:
				tempValue = int(tempValue)
			except (TypeError, ValueError):
				try:
					tempValue = float(tempValue)
				except (TypeError, ValueError):
					if checkbool(tempValue) == 1:
						tempValue = True
					elif checkbool(tempValue) == 0:
						tempValue = False
		excel_dict[sheet.cell(row = 1, column = j).value] = tempValue
	excel_list.append(excel_dict)
	excel_dict = {}

sheet_name = sheet.title

with open(sheet_name + ".json", mode='a', encoding='utf-8') as f:
	jsonvalue = json.dumps(excel_list, ensure_ascii=False, sort_keys=True, indent=4)
	f.write(jsonvalue)